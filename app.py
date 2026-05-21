"""
Flask app — Análisis COD (Cash On Delivery).

Endpoints:
  GET  /                              → dashboard
  POST /reports/cod/scrape-dropi      → arranca scrape+análisis (worker thread)
  GET  /reports/cod/scrape-dropi/stream  → SSE de progreso/result
  POST /reports/cod/analyze-upload    → fallback: subir Excel a mano
  GET  /reports/cod/healthcheck       → diagnóstico Meta/Shopify
"""
from __future__ import annotations

import json
import os
import queue
import sys
import threading
import traceback
from datetime import date, datetime

from io import BytesIO

from flask import (Flask, Response, jsonify, redirect, render_template,
                   request, send_file, stream_with_context)

import config
from models_reports import CodConfig

# Consola Windows: evitar UnicodeEncodeError (cp1252) al loguear.
for _s in (sys.stdout, sys.stderr):
    try:
        _s.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

app = Flask(__name__)
# Recarga templates al vuelo cuando se editan (sin reiniciar el server).
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.jinja_env.auto_reload = True

# ── Auth por token (solo cuando se despliega público, ej. Railway) ──
# Si APP_ACCESS_TOKEN está seteado: cualquier request requiere el token via
# ?token=... la primera vez (después queda en cookie por 1 año).
# Si está vacío (local dev): la app no exige nada.
APP_ACCESS_TOKEN = os.getenv("APP_ACCESS_TOKEN", "").strip()


@app.before_request
def _check_token():
    if not APP_ACCESS_TOKEN:
        return None  # Modo local sin auth
    if request.path == "/favicon.ico":
        return None
    if request.cookies.get("cod_auth") == APP_ACCESS_TOKEN:
        return None
    if request.args.get("token", "") == APP_ACCESS_TOKEN:
        # Set cookie + redirect a la misma URL sin el token visible
        clean = request.path
        # Preservar otros query params
        other = {k: v for k, v in request.args.items() if k != "token"}
        if other:
            from urllib.parse import urlencode
            clean += "?" + urlencode(other)
        resp = redirect(clean)
        is_https = (request.is_secure
                    or request.headers.get("X-Forwarded-Proto") == "https")
        resp.set_cookie("cod_auth", APP_ACCESS_TOKEN,
                        max_age=60 * 60 * 24 * 365,
                        httponly=True, samesite="Lax", secure=is_https)
        return resp
    return (
        "<html><body style='background:#0f1419;color:#e6edf3;"
        "font:14px/1.5 system-ui;padding:40px;max-width:520px;margin:auto'>"
        "<h2>Acceso restringido</h2>"
        "<p>Esta app requiere un token. Agrega <code>?token=...</code> al "
        "final de la URL (la primera vez; después queda guardado en cookie).</p>"
        "</body></html>", 401)

# Single-user: una cola global + lock de "ocupado". Para multiusuario,
# reemplazar por dict {job_id: Queue}.
_scrape_queue: "queue.Queue[tuple[str, str]]" = queue.Queue()
_busy = threading.Lock()

# Cache del último análisis (para /recompute con overrides nuevos sin
# re-scrapear Dropi ni re-subir el CSV).
_last_inputs: dict = {}
_last_result: dict = {}  # último result.to_dict() para descargas (Excel, etc.)


def _parse_date(s: str) -> date:
    return datetime.strptime(str(s).strip(), "%Y-%m-%d").date()


def _cfg_from(src: dict) -> tuple[CodConfig, bool]:
    """Construye CodConfig desde el form/json. Devuelve (cfg, use_real)."""
    def f(k, d):
        try:
            return float(src.get(k))
        except (TypeError, ValueError):
            return d
    cfg = CodConfig(
        tasa_confirmacion=f("tasa_confirmacion", 0.75),
        tasa_entrega=f("tasa_entrega", 0.70),
        costo_envio=f("costo_envio", 3.0),
        gastos_operativos=f("gastos_operativos", 0.0),
    )
    use_real = str(src.get("use_real", "true")).lower() not in ("false", "0", "no")
    return cfg, use_real


def _run_cod_analysis_from_bytes(file_bytes, account_id, gastos_op,
                                 cfg, use_real, date_from, date_to,
                                 manual_spend=None, meta_csv_bytes=None,
                                 meta_csv_filename=None,
                                 progress_cb=None) -> dict:
    from sources.cod_analysis import analyze_bytes
    cfg = cfg or CodConfig(gastos_operativos=float(gastos_op or 0))
    return analyze_bytes(
        file_bytes, config=cfg, date_from=date_from, date_to=date_to,
        use_real=use_real, do_attribution=True, manual_spend_text=manual_spend,
        meta_csv_bytes=meta_csv_bytes, meta_csv_filename=meta_csv_filename,
        source_filename="dropi_scrape.xlsx", progress_cb=progress_cb,
    )


@app.route("/")
def index():
    return render_template("dashboard.html")


@app.route("/reports/cod/scrape-dropi", methods=["POST"])
def scrape_dropi_start():
    if _busy.locked():
        return jsonify({"ok": False,
                        "error": "Ya hay un scrape en curso."}), 409
    # Acepta JSON o multipart (multipart cuando se sube CSV de Meta)
    if request.is_json:
        src = request.get_json(force=True, silent=True) or {}
    else:
        src = request.form
    try:
        date_from = _parse_date(src.get("date_from") or "")
        date_to = _parse_date(src.get("date_to") or "")
    except (TypeError, ValueError):
        return jsonify({"ok": False,
                        "error": "date_from / date_to inválidos (YYYY-MM-DD)."}), 400
    account_id = src.get("account_id")
    cfg, use_real = _cfg_from(src)
    cfg.gastos_operativos = float(src.get("gastos_operativos", 0) or 0)
    manual_spend = src.get("manual_spend") or None

    # Archivo opcional: CSV/XLSX exportado del Ads Manager
    meta_csv_bytes = None
    meta_csv_filename = None
    if not request.is_json:
        f = request.files.get("meta_csv")
        if f and f.filename:
            meta_csv_bytes = f.read()
            meta_csv_filename = f.filename

    job_id = datetime.now().strftime("%Y%m%d%H%M%S")

    # Drenar cola de un job anterior
    while not _scrape_queue.empty():
        try:
            _scrape_queue.get_nowait()
        except queue.Empty:
            break

    def worker():
        _busy.acquire()

        def progress_cb(msg: str):
            _scrape_queue.put(("progress", str(msg).replace("\n", " | ")))

        try:
            from sources.dropi_report import download_orders_report
            path = download_orders_report(date_from, date_to,
                                          progress_cb=progress_cb)
            with open(path, "rb") as fh:
                file_bytes = fh.read()
            progress_cb(f"Procesando {path.name} con el pipeline COD…")
            # Cachear inputs para recompute con overrides
            _last_inputs.clear()
            _last_inputs.update({
                "file_bytes": file_bytes, "cfg": cfg, "use_real": use_real,
                "date_from": date_from, "date_to": date_to,
                "manual_spend": manual_spend,
                "meta_csv_bytes": meta_csv_bytes,
                "meta_csv_filename": meta_csv_filename,
                "source_filename": path.name,
            })
            result = _run_cod_analysis_from_bytes(
                file_bytes, account_id, cfg.gastos_operativos, cfg,
                use_real, date_from, date_to, manual_spend=manual_spend,
                meta_csv_bytes=meta_csv_bytes,
                meta_csv_filename=meta_csv_filename,
                progress_cb=progress_cb)
            _maybe_upload_to_drive(result, progress_cb=progress_cb)
            _last_result.clear(); _last_result.update(result)
            _scrape_queue.put(("result", json.dumps(result, default=str)))
        except Exception as e:
            _scrape_queue.put(("progress",
                               traceback.format_exc().replace("\n", " | ")))
            _scrape_queue.put(("error", str(e)))
        finally:
            _scrape_queue.put(("done", ""))
            _busy.release()

    threading.Thread(target=worker, daemon=True).start()
    return jsonify({"ok": True, "job_id": job_id}), 202


@app.route("/reports/cod/scrape-dropi/stream")
def scrape_dropi_stream():
    @stream_with_context
    def gen():
        yield "retry: 3000\n\n"
        while True:
            try:
                event, data = _scrape_queue.get(timeout=30)
            except queue.Empty:
                yield ": keep-alive\n\n"
                continue
            yield f"event: {event}\ndata: {data}\n\n"
            if event == "done":
                break

    return Response(gen(), mimetype="text/event-stream", headers={
        "Cache-Control": "no-cache",
        "X-Accel-Buffering": "no",
        "Connection": "keep-alive",
    })


@app.route("/reports/cod/analyze-upload", methods=["POST"])
def analyze_upload():
    if "file" not in request.files:
        return jsonify({"ok": False, "error": "Falta el archivo .xlsx"}), 400
    f = request.files["file"]
    if not f.filename.lower().endswith(".xlsx"):
        return jsonify({"ok": False, "error": "Subí un .xlsx"}), 400
    cfg, use_real = _cfg_from(request.form)
    date_from = date_to = None
    try:
        if request.form.get("date_from"):
            date_from = _parse_date(request.form["date_from"])
        if request.form.get("date_to"):
            date_to = _parse_date(request.form["date_to"])
    except ValueError:
        pass
    try:
        from sources.cod_analysis import analyze_bytes
        file_bytes = f.read()
        manual_spend = request.form.get("manual_spend") or None
        # Cachear para recompute (sin re-subir)
        _last_inputs.clear()
        _last_inputs.update({
            "file_bytes": file_bytes, "cfg": cfg, "use_real": use_real,
            "date_from": date_from, "date_to": date_to,
            "manual_spend": manual_spend,
            "meta_csv_bytes": None, "meta_csv_filename": None,
            "source_filename": f.filename,
        })
        result = analyze_bytes(
            file_bytes, config=cfg, date_from=date_from, date_to=date_to,
            use_real=use_real, do_attribution=True,
            manual_spend_text=manual_spend,
            source_filename=f.filename)
        _maybe_upload_to_drive(result)
        _last_result.clear(); _last_result.update(result)
        return jsonify({"ok": True, "result": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e),
                        "trace": traceback.format_exc()}), 500


@app.route("/reports/cod/override-campaign", methods=["POST"])
def override_campaign():
    """Guarda/quita un mapeo manual campaña → producto. No re-corre el análisis."""
    from sources import campaign_overrides
    body = request.get_json(force=True, silent=True) or {}
    cname = (body.get("campaign") or "").strip()
    pid = (body.get("product_id") or "").strip() or None
    if not cname:
        return jsonify({"ok": False, "error": "Falta 'campaign'"}), 400
    campaign_overrides.save_override(cname, pid)
    return jsonify({"ok": True, "campaign": cname, "product_id": pid})


@app.route("/reports/cod/recompute", methods=["POST"])
def recompute():
    """Recalcula análisis con los inputs del último run + overrides actuales."""
    if not _last_inputs:
        return jsonify({"ok": False, "error": "No hay análisis previo. "
                        "Corré primero el scraper o subí un Excel."}), 400
    try:
        from sources.cod_analysis import analyze_bytes
        result = analyze_bytes(
            _last_inputs["file_bytes"],
            config=_last_inputs["cfg"],
            date_from=_last_inputs["date_from"],
            date_to=_last_inputs["date_to"],
            use_real=_last_inputs["use_real"],
            do_attribution=True,
            manual_spend_text=_last_inputs.get("manual_spend"),
            meta_csv_bytes=_last_inputs.get("meta_csv_bytes"),
            meta_csv_filename=_last_inputs.get("meta_csv_filename"),
            source_filename=_last_inputs.get("source_filename", ""))
        _last_result.clear(); _last_result.update(result)
        return jsonify({"ok": True, "result": result})
    except Exception as e:
        return jsonify({"ok": False, "error": str(e),
                        "trace": traceback.format_exc()}), 500


def _build_stale_excel_bytes(rows: list) -> bytes:
    """Genera el .xlsx de pendientes (usado por descarga y por subida a Drive)."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    ws = wb.active
    ws.title = "Pendientes guía >3d"
    if not rows:
        ws.append(["No hay pedidos pendientes con guía >3 días."])
    else:
        headers = list(rows[0].keys())
        ws.append(headers)
        hdr_fill = PatternFill("solid", fgColor="1F3A5F")
        hdr_font = Font(bold=True, color="FFFFFF")
        for col_idx, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col_idx)
            c.fill = hdr_fill; c.font = hdr_font
            c.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append([r.get(k, "") for k in headers])
        for col_idx, h in enumerate(headers, 1):
            mx = max(len(str(h)),
                     max((len(str(r.get(h, ""))) for r in rows), default=10))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max(mx + 2, 12), 45)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _stale_filename() -> str:
    from datetime import datetime as _dt
    return (f"pendientes_guia_mas_3d_"
            f"{_dt.now().strftime('%Y%m%d_%H%M')}.xlsx")


def _maybe_upload_to_drive(result: dict, progress_cb=None) -> None:
    """Si Drive está configurado y hay pendientes, sube Excel y guarda
    drive_link en result. Mutación in-place."""
    rows = result.get("stale_pending_orders") or []
    if not rows:
        return
    from sources import drive_upload
    if not drive_upload.is_configured():
        return
    if progress_cb:
        progress_cb("Subiendo Excel de pendientes a Drive…")
    xlsx = _build_stale_excel_bytes(rows)
    url, err = drive_upload.upload_file(xlsx, _stale_filename())
    if err:
        if progress_cb:
            progress_cb(f"  ⚠ Drive upload falló: {err}")
        result.setdefault("warnings", []).append(f"Drive upload: {err}")
    else:
        if progress_cb:
            progress_cb(f"  ✓ Excel en Drive: {url}")
        result["drive_link"] = url


@app.route("/reports/cod/stale-pending.xlsx")
def download_stale_pending():
    """Descarga local del Excel de pendientes."""
    rows = (_last_result.get("stale_pending_orders") or [])
    data = _build_stale_excel_bytes(rows)
    return send_file(BytesIO(data),
        mimetype=("application/vnd.openxmlformats-officedocument."
                  "spreadsheetml.sheet"),
        as_attachment=True, download_name=_stale_filename())


@app.route("/reports/cod/upload-stale-now", methods=["POST"])
def upload_stale_now():
    """Sube manualmente a Drive el Excel del último análisis."""
    if not _last_result.get("stale_pending_orders"):
        return jsonify({"ok": False,
                        "error": "No hay pendientes en el último análisis."}), 400
    from sources import drive_upload
    if not drive_upload.is_configured():
        return jsonify({"ok": False,
                        "error": "Drive no configurado (SA + folder id)."}), 400
    xlsx = _build_stale_excel_bytes(_last_result["stale_pending_orders"])
    url, err = drive_upload.upload_file(xlsx, _stale_filename())
    if err:
        return jsonify({"ok": False, "error": err}), 500
    _last_result["drive_link"] = url
    return jsonify({"ok": True, "drive_link": url})


@app.route("/reports/cod/healthcheck")
def healthcheck():
    from sources.ad_spend import check_meta_access
    from sources.shopify_attribution import check_shopify_access
    from sources.meta_spend_email import check_gmail_access
    from sources.drive_upload import check_drive_access, get_sa_email
    m_ok, m_msg = check_meta_access()
    s_ok, s_msg = check_shopify_access()
    g_ok, g_msg = check_gmail_access()
    d_ok, d_msg = check_drive_access()
    return jsonify({
        "dropi_email_set": bool(config.DROPI_EMAIL),
        "dropi_password_set": bool(config.DROPI_PASSWORD),
        "meta": {"ok": m_ok, "msg": m_msg,
                 "ad_account": config.META_AD_ACCOUNT_ID_COD},
        "shopify": {"ok": s_ok, "msg": s_msg},
        "gmail": {"ok": g_ok, "msg": g_msg,
                  "address": config.GMAIL_ADDRESS,
                  "subject": config.GMAIL_REPORT_SUBJECT},
        "drive": {"ok": d_ok, "msg": d_msg,
                  "sa_email": get_sa_email(),
                  "folder_id": config.GOOGLE_DRIVE_FOLDER_ID},
        "campaign_regex_set": bool(config.META_CAMPAIGN_PRODUCT_REGEX),
    })


if __name__ == "__main__":
    # HOST=0.0.0.0 en Railway (Docker), 127.0.0.1 por defecto en local.
    host = os.getenv("HOST", "127.0.0.1")
    print(f"COD Analysis · http://{host}:{config.PORT}")
    app.run(host=host, port=config.PORT, threaded=True, debug=False)
