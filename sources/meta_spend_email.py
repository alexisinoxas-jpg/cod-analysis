"""
Lectura automática del reporte programado de Meta Ads desde Gmail (IMAP).

Cómo funciona:
  1. Conecta a Gmail vía IMAP con contraseña de aplicación (no OAuth).
  2. Busca el correo más reciente del informe programado (por asunto/sender).
  3. Baja el adjunto CSV/XLSX, lo parsea (campaña + gasto, opcional por día).
  4. Agrega gasto por nombre de campaña, atribuye a producto con el mismo
     regex que ya usamos, y devuelve un MetaSpendResult listo para el pipeline.

No usa Meta API → no toca el portal de developers (compatible con la sanción).
"""
from __future__ import annotations

import csv
import email
import imaplib
import io
import re
import unicodedata
from datetime import date, datetime
from email.header import decode_header

import config
from sources.ad_spend import (CampaignSpend, MetaSpendResult, _compiled_regex,
                              _parse_campaign_name, normalize)


# ── Utilidades de parsing ───────────────────────────────────────────────
def _h(s) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s.strip().lower())


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_day(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _decode(s: str) -> str:
    parts = decode_header(s or "")
    out = ""
    for txt, enc in parts:
        if isinstance(txt, bytes):
            try:
                out += txt.decode(enc or "utf-8", errors="replace")
            except LookupError:
                out += txt.decode("utf-8", errors="replace")
        else:
            out += txt
    return out


# ── Detección de columnas en el CSV/XLSX del informe ────────────────────
def _find_header_row(rows: list[list]) -> int:
    """Devuelve el índice de la fila que parece ser el header.
    Ads Manager a veces antepone filas de metadata. Buscamos la primera
    fila que tiene a la vez nombre de campaña y monto/gasto."""
    for i, r in enumerate(rows[:15]):
        cells = [_h(c) for c in r]
        joined = " | ".join(cells)
        if (("campa" in joined or "campaign" in joined)
                and ("gast" in joined or "spend" in joined or "amount" in joined
                     or "importe" in joined)):
            return i
    return 0


def _column_indices(header: list[str]):
    name_idx = spend_idx = day_idx = purch_idx = conv_idx = None
    for i, h in enumerate(header):
        hh = _h(h)
        if name_idx is None and ("nombre de la campa" in hh
                                  or "campaign name" in hh
                                  or hh == "campaña" or hh == "campaign"):
            name_idx = i
        if spend_idx is None and ("importe gastado" in hh
                                   or "amount spent" in hh
                                   or hh.startswith("gast")
                                   or "spend" in hh):
            spend_idx = i
        if day_idx is None and (hh == "día" or hh == "dia" or hh == "day"
                                 or hh == "fecha" or hh == "date"):
            day_idx = i
        # "Compras" o "Resultados" (cuenta de compras del píxel)
        if purch_idx is None and ("compras" == hh or "purchases" == hh
                                   or "resultados" == hh):
            purch_idx = i
        # "Valor de conversión de compras" / "Purchase value" / "ROAS"
        # Importante: detectar VALOR de conversión, no la columna "% de…"
        if conv_idx is None and (
                "valor de conversion" in hh
                or "valor de conversi" in hh
                or "purchase conversion value" in hh
                or "purchase value" in hh
                or "valor de las compras" in hh):
            conv_idx = i
    return name_idx, spend_idx, day_idx, purch_idx, conv_idx


# ── Aggregación spend por campaña ───────────────────────────────────────
def _aggregate(rows_iter, name_idx, spend_idx, day_idx, purch_idx, conv_idx,
               date_from: date | None, date_to: date | None,
               source: str) -> MetaSpendResult:
    res = MetaSpendResult()
    rx = _compiled_regex()
    # Por campaña: spend, purchases (cuenta), conv_value (valor de compra)
    by_cname: dict[str, dict] = {}

    for r in rows_iter:
        if name_idx is None or spend_idx is None:
            break
        if not r or name_idx >= len(r) or spend_idx >= len(r):
            continue
        cname = str(r[name_idx] or "").strip()
        if not cname or _h(cname) in ("total", "totales", "totals"):
            continue
        if day_idx is not None and day_idx < len(r):
            d = _parse_day(r[day_idx])
            if d is None:
                continue
            if date_from and d < date_from:
                continue
            if date_to and d > date_to:
                continue
        spend = _to_float(r[spend_idx])
        if spend <= 0:
            continue
        purchases = (int(round(_to_float(r[purch_idx])))
                     if purch_idx is not None and purch_idx < len(r) else 0)
        conv_value = (_to_float(r[conv_idx])
                      if conv_idx is not None and conv_idx < len(r) else 0.0)
        d = by_cname.setdefault(
            cname, {"spend": 0.0, "purchases": 0, "conv_value": 0.0})
        d["spend"] += spend
        d["purchases"] += purchases
        d["conv_value"] += conv_value

    for cname, dat in by_cname.items():
        key, pname = _parse_campaign_name(cname, rx)
        res.campaigns.append(CampaignSpend(
            campaign_id="", campaign_name=cname, spend=dat["spend"],
            product_key=key, product_name=pname,
            matched=bool(key or pname),
            purchases=dat["purchases"], conv_value=dat["conv_value"]))
        res.total_spend += dat["spend"]
        if key:
            res.by_key[key] = res.by_key.get(key, 0.0) + dat["spend"]
        if pname:
            nk = normalize(pname)
            res.by_name[nk] = res.by_name.get(nk, 0.0) + dat["spend"]
        if not (key or pname):
            res.warnings.append(
                f"Campaña sin patrón reconocible: «{cname}» "
                f"(${dat['spend']:,.0f} sin atribuir)")

    if res.campaigns:
        res.warnings.append(
            f"Reporte Meta procesado: {len(res.campaigns)} campañas, "
            f"${res.total_spend:,.0f} total (origen: {source}).")
    return res


def _parse_csv_bytes(data: bytes) -> tuple[list[list[str]], str]:
    text = None
    for enc in ("utf-8", "utf-8-sig", "latin-1", "cp1252"):
        try:
            text = data.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = data.decode("utf-8", errors="replace")
    # Sniff dialecto
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect=dialect))
    return rows, "csv"


def _parse_xlsx_bytes(data: bytes) -> tuple[list[list], str]:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True,
                                data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    return rows, "xlsx"


def parse_report_bytes(filename: str, data: bytes,
                       date_from: date | None = None,
                       date_to: date | None = None) -> MetaSpendResult:
    """Parsea CSV o XLSX → MetaSpendResult, agregado por nombre de campaña."""
    fn = (filename or "").lower()
    if fn.endswith(".xlsx") or fn.endswith(".xlsm"):
        rows, kind = _parse_xlsx_bytes(data)
    else:
        rows, kind = _parse_csv_bytes(data)

    if not rows:
        res = MetaSpendResult()
        res.warnings.append(f"Reporte vacío ({filename})")
        return res

    hr = _find_header_row(rows)
    header = rows[hr]
    name_idx, spend_idx, day_idx, purch_idx, conv_idx = _column_indices(header)
    if name_idx is None or spend_idx is None:
        res = MetaSpendResult()
        res.warnings.append(
            f"No reconozco las columnas del reporte ({filename}). "
            f"Header detectado: {header[:6]!r}")
        return res

    return _aggregate(iter(rows[hr + 1:]), name_idx, spend_idx, day_idx,
                      purch_idx, conv_idx,
                      date_from, date_to,
                      source=f"{filename} ({kind})")


# ── IMAP: traer el último correo del informe ────────────────────────────
def fetch_latest_report_attachment(progress_cb=None
                                   ) -> tuple[str | None, bytes | None,
                                              str | None]:
    """Conecta a Gmail por IMAP y devuelve (filename, bytes, error_msg)."""
    def n(msg):
        if progress_cb:
            progress_cb(msg)

    user, pw = config.GMAIL_ADDRESS, config.GMAIL_APP_PASSWORD
    if not user or not pw:
        return None, None, ("Faltan GMAIL_ADDRESS / GMAIL_APP_PASSWORD "
                            "en .env (contraseña de aplicación).")

    subject = config.GMAIL_REPORT_SUBJECT
    sender = config.GMAIL_REPORT_SENDER

    n(f"Conectando a Gmail (IMAP) como {user}…")
    try:
        m = imaplib.IMAP4_SSL("imap.gmail.com", 993)
    except OSError as e:
        return None, None, f"No pude conectar a imap.gmail.com: {e}"
    try:
        m.login(user, pw)
    except imaplib.IMAP4.error as e:
        return None, None, (f"Login IMAP falló: {e}. ¿Contraseña de "
                            f"aplicación correcta? ¿2FA activado?")
    try:
        m.select("INBOX", readonly=True)
        # IMAP search no soporta bien caracteres no-ASCII en el criterio
        # (ej. 'ñ' en el subject). Estrategia: buscar por FROM (ASCII) y
        # filtrar el subject del lado de Python sobre los últimos N msgs.
        n(f"Buscando correos: FROM {sender}")
        typ, data = m.search(None, 'FROM', f'"{sender}"')
        ids = (data[0] or b"").split()
        if not ids:
            return None, None, (
                f"No encontré correos de {sender} en INBOX. ¿Ya programaste "
                f"el informe y llegó el primero?")

        latest_id = None
        subj_needle = (subject or "").lower().strip()
        if subj_needle:
            # Revisar headers de los últimos 25 msgs (los más recientes primero)
            for mid in reversed(ids[-25:]):
                typ, hdata = m.fetch(
                    mid, "(BODY.PEEK[HEADER.FIELDS (SUBJECT)])")
                if not hdata or not hdata[0]:
                    continue
                hbytes = hdata[0][1] if isinstance(hdata[0], tuple) else b""
                try:
                    raw_hdr = hbytes.decode("utf-8", errors="replace")
                except Exception:
                    raw_hdr = ""
                msub = re.search(r"Subject:\s*(.*)", raw_hdr,
                                  re.IGNORECASE | re.DOTALL)
                actual = _decode(msub.group(1).strip()) if msub else ""
                if subj_needle in actual.lower():
                    latest_id = mid
                    break
            if latest_id is None:
                return None, None, (
                    f"No encontré correos cuyo asunto contenga «{subject}» "
                    f"entre los últimos 25 de {sender}.")
        else:
            latest_id = ids[-1]

        n(f"Bajando el correo (id {latest_id.decode()})…")
        typ, msg_data = m.fetch(latest_id, "(RFC822)")
        raw = msg_data[0][1]
    finally:
        try:
            m.logout()
        except Exception:
            pass

    msg = email.message_from_bytes(raw)
    for part in msg.walk():
        if part.get_content_disposition() != "attachment":
            continue
        fn = _decode(part.get_filename() or "")
        if not fn:
            continue
        if fn.lower().endswith((".csv", ".xlsx", ".xlsm")):
            payload = part.get_payload(decode=True)
            return fn, payload, None

    return None, None, ("Correo encontrado pero sin adjunto CSV/XLSX. "
                        "Revisá que el reporte esté programado en formato "
                        "CSV.")


def fetch_meta_spend_from_email(date_from: date | None = None,
                                date_to: date | None = None,
                                progress_cb=None) -> MetaSpendResult:
    """API pública: trae spend Meta del último correo programado."""
    fn, data, err = fetch_latest_report_attachment(progress_cb=progress_cb)
    if err:
        res = MetaSpendResult()
        res.warnings.append(err)
        return res
    if progress_cb:
        progress_cb(f"Parseando {fn} ({len(data)//1024} KB)…")
    return parse_report_bytes(fn, data, date_from=date_from, date_to=date_to)


def check_gmail_access() -> tuple[bool, str]:
    """Healthcheck: ¿IMAP login OK?"""
    if not config.GMAIL_ADDRESS or not config.GMAIL_APP_PASSWORD:
        return False, ("Sin GMAIL_ADDRESS / GMAIL_APP_PASSWORD — modo "
                       "AUTO-EMAIL no configurado (se usa pegado manual).")
    try:
        m = imaplib.IMAP4_SSL("imap.gmail.com", 993)
        m.login(config.GMAIL_ADDRESS, config.GMAIL_APP_PASSWORD)
        m.logout()
        return True, f"OK · IMAP login para {config.GMAIL_ADDRESS}"
    except imaplib.IMAP4.error as e:
        return False, f"IMAP login falló: {e}"
    except OSError as e:
        return False, f"Red: {e}"
