"""
Pipeline COD: Excel "Órdenes con Productos (un producto por fila)" de Dropi
→ list[ProductReportRow] + totales.

Usa tasas REALES de ESTATUS (no estimadas). Los sliders de CodConfig
funcionan como override de simulación: si `use_real=False` se proyecta
con las fórmulas del brief sobre los brutos.

Atribución FB/TikTok: vía Shopify (ID DE ORDEN DE TIENDA → utm_source).
Spend Meta por producto: vía ad_spend (convención de nombres de campaña).
"""
from __future__ import annotations

import io
import re
import unicodedata
from collections import defaultdict
from datetime import date, datetime

import openpyxl

import config as appcfg
from models_reports import CodConfig, CodAnalysisResult, ProductReportRow
from sources import ad_spend, shopify_attribution

# ── Clasificación de ESTATUS ─────────────────────────────────
DELIVERED = {"ENTREGADO"}
RETURNED = {"DEVOLUCION", "EN DEVOLUCION", "DEVOLUCION A REMITENTE"}
CANCELLED = {"CANCELADO"}
# cualquier otro ESTATUS → PENDING (en tránsito, sin resolver)


def _norm(s) -> str:
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^A-Za-z0-9]+", " ", s).strip().upper()
    return s


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(" ", "")
    # formato chileno "1.234,56" → "1234.56"
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _to_int(v) -> int:
    return int(round(_to_float(v)))


def _parse_fecha(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v or "").strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%m-%y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# Headers esperados (normalizados) → nombre lógico
_COLMAP = {
    "ID": "order_id",
    "FECHA": "fecha",
    "ESTATUS": "estatus",
    "TOTAL DE LA ORDEN": "order_total",
    "GANANCIA": "ganancia",
    "PRECIO FLETE": "flete",
    "COSTO DEVOLUCION FLETE": "flete_devol",
    "PRECIO PROVEEDOR": "prov_unit",
    "PRECIO PROVEEDOR X CANTIDAD": "prov_total",
    "PRODUCTO ID": "product_id",
    "SKU": "sku",
    "PRODUCTO": "product_title",
    "CANTIDAD": "cantidad",
    "ID DE ORDEN DE TIENDA": "shop_order_id",
    "NUMERO DE PEDIDO DE TIENDA": "shop_order_num",
    "FECHA GUIA GENERADA": "fecha_guia",
    # Para exportar detalle de pedidos pendientes
    "NUMERO GUIA": "numero_guia",
    "NOMBRE CLIENTE": "cliente",
    "TELEFONO": "telefono",
    "CIUDAD DESTINO": "ciudad",
    "DEPARTAMENTO DESTINO": "departamento",
    "DIRECCION": "direccion",
    "TRANSPORTADORA": "transportadora",
    "NOVEDAD": "novedad",
    "ULTIMO MOVIMIENTO": "ultimo_mov",
    "FECHA DE ULTIMO MOVIMIENTO": "fecha_ultimo_mov",
    "NOTAS": "notas",
}

STALE_PENDING_DAYS = 3  # umbral para alertar pendientes "varados"

# La alerta de "varado" SOLO aplica a estos estatus (pre-envío). Si la orden
# ya está EN TRÁNSITO / EN REPARTO / EN RETIRO / NOVEDAD, está moviéndose
# (o tiene su propia categoría de problema) y no entra en esta alerta.
EARLY_STUCK_STATUSES = {
    "GUIA GENERADA", "GUIA_GENERADA",     # guía emitida, transportadora aún no la levantó
    "PENDIENTE", "PENDING",
}
_REQUIRED = {"estatus", "order_total", "prov_total", "product_id",
             "product_title", "cantidad"}


def _resolve_columns(header: list) -> tuple[dict, list[str]]:
    """normalized header → idx; devuelve (logical→idx, warnings)."""
    norm_to_idx = {}
    for i, h in enumerate(header):
        norm_to_idx.setdefault(_norm(h), i)
    logical = {}
    warnings = []
    for exp_norm, logical_name in _COLMAP.items():
        if exp_norm in norm_to_idx:
            logical[logical_name] = norm_to_idx[exp_norm]
    missing = _REQUIRED - set(logical)
    if missing:
        warnings.append(
            "Columnas requeridas no encontradas: "
            + ", ".join(sorted(missing))
            + ". ¿Es el reporte 'Órdenes con Productos (un producto por fila)'?"
        )
    return logical, warnings


def _status_bucket(estatus: str) -> str:
    e = _norm(estatus)
    if e in DELIVERED:
        return "delivered"
    if e in RETURNED:
        return "returned"
    if e in CANCELLED:
        return "cancelled"
    return "pending"


def _safe_div(a: float, b: float) -> float:
    return a / b if b else 0.0


# ─────────────────────────────────────────────────────────────
# API pública
# ─────────────────────────────────────────────────────────────
def analyze_bytes(
    file_bytes: bytes,
    *,
    config: CodConfig | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    use_real: bool = True,
    do_attribution: bool = True,
    manual_spend_text: str | None = None,
    meta_csv_bytes: bytes | None = None,
    meta_csv_filename: str | None = None,
    source_filename: str = "",
    progress_cb=None,
) -> dict:
    return analyze_excel(
        io.BytesIO(file_bytes), config=config, date_from=date_from,
        date_to=date_to, use_real=use_real, do_attribution=do_attribution,
        manual_spend_text=manual_spend_text,
        meta_csv_bytes=meta_csv_bytes,
        meta_csv_filename=meta_csv_filename,
        source_filename=source_filename, progress_cb=progress_cb,
    ).to_dict()


def analyze_excel(
    source,
    *,
    config: CodConfig | None = None,
    date_from: date | None = None,
    date_to: date | None = None,
    use_real: bool = True,
    do_attribution: bool = True,
    manual_spend_text: str | None = None,
    meta_csv_bytes: bytes | None = None,
    meta_csv_filename: str | None = None,
    source_filename: str = "",
    progress_cb=None,
) -> CodAnalysisResult:
    def n(msg):
        if progress_cb:
            progress_cb(msg)

    cfg = config or CodConfig()
    result = CodAnalysisResult(config=cfg, source_filename=source_filename)

    wb = openpyxl.load_workbook(source, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    H = len(header)
    cols, warns = _resolve_columns(header)
    result.warnings.extend(warns)

    def cell(row, name):
        idx = cols.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    rows = []
    dates_seen = []
    for raw in it:
        raw = list(raw) + [None] * (H - len(raw))
        if all(v in (None, "") for v in raw):
            continue
        f = _parse_fecha(cell(raw, "fecha"))
        if f:
            dates_seen.append(f)
        rows.append(raw)

    if dates_seen:
        if not date_from:
            date_from = min(dates_seen)
        if not date_to:
            date_to = max(dates_seen)
    n(f"Filas de datos: {len(rows)} · rango {date_from} → {date_to}")

    # ── Funnel Shopify (todas las órdenes del rango con line items) ──
    # Sirve para 2 cosas: 1) atribución FB/TikTok por orden (utm_source),
    # 2) ver cuántas órdenes existieron en Shopify (haya o no entrado a Dropi).
    shop = shopify_attribution.ShopifyOrdersResult()
    if do_attribution and date_from and date_to:
        n("Trayendo todas las órdenes Shopify del rango (paginado)…")
        shop = shopify_attribution.fetch_orders_in_range(
            date_from, date_to, progress_cb=n)
        result.warnings.extend(shop.warnings)
        n(f"  Shopify: {shop.fetched} órdenes en el rango")
    elif do_attribution:
        result.warnings.append(
            "Falta date_from/date_to — Shopify (atribución y funnel) deshabilitado.")

    # ── Spend Meta por producto (4 fuentes, en orden de prioridad) ──
    #   1. CSV/XLSX subido a mano (override explícito, ej. "Exportar" del AM).
    #   2. Pegado manual del Ads Manager.
    #   3. Modo AUTO: lectura del reporte programado vía Gmail (IMAP).
    #   4. API Meta (sólo si META_API_ENABLED y hay token vivo).
    meta = ad_spend.MetaSpendResult()
    if meta_csv_bytes:
        from sources import meta_spend_email
        n(f"Parseando CSV/XLSX del Ads Manager: {meta_csv_filename or 'archivo'}…")
        meta = meta_spend_email.parse_report_bytes(
            meta_csv_filename or "manual.csv", meta_csv_bytes,
            date_from=date_from, date_to=date_to)
        result.warnings.extend(meta.warnings)
        n(f"  Archivo: ${meta.total_spend:,.0f} en "
          f"{len(meta.campaigns)} campañas")
    elif manual_spend_text and manual_spend_text.strip():
        n("Procesando gasto Meta MANUAL pegado del Ads Manager…")
        meta = ad_spend.parse_manual_spend(manual_spend_text)
        result.warnings.extend(meta.warnings)
        n(f"  Manual: ${meta.total_spend:,.0f} en "
          f"{len(meta.campaigns)} campañas")
    elif appcfg.GMAIL_ADDRESS and appcfg.GMAIL_APP_PASSWORD:
        from sources import meta_spend_email
        n("Trayendo gasto Meta del correo programado (Gmail/IMAP)…")
        meta = meta_spend_email.fetch_meta_spend_from_email(
            date_from=date_from, date_to=date_to, progress_cb=n)
        result.warnings.extend(meta.warnings)
        if meta.total_spend > 0:
            n(f"  Auto-email: ${meta.total_spend:,.0f} en "
              f"{len(meta.campaigns)} campañas")
    elif date_from and date_to:
        n("Trayendo spend Meta de la cuenta COD (API)…")
        meta = ad_spend.fetch_meta_spend(date_from, date_to)
        result.warnings.extend(meta.warnings)
        n(f"  Meta: ${meta.total_spend:,.0f} en {len(meta.campaigns)} campañas")

    # ── Agrupar por orden (para no doble-contar montos de orden) ──
    orders = defaultdict(list)
    for r in rows:
        orders[cell(r, "order_id")].append(r)

    # ── Acumular por producto ──
    agg: dict[str, dict] = {}

    def bucket(pid):
        if pid not in agg:
            agg[pid] = dict(
                product_id=pid, product_title="", sku="",
                units=0, revenue_gross=0.0, cogs_gross=0.0, orders=0,
                d=0, c=0, ret=0, pend=0, pend_stale=0,
                rev_real=0.0, cogs_real=0.0, ship=0.0,
                of=0, ot=0, ou=0, rf=0.0, rt=0.0, ru=0.0,
                uf=0, ut=0, uu=0,
                prov_unit_sum=0.0, prov_unit_n=0,
            )
        return agg[pid]

    today = datetime.now().date()
    stale_orders_collected: list[dict] = []

    for oid, olines in orders.items():
        n_lines = len(olines)
        line_cogs = [_to_float(cell(l, "prov_total")) for l in olines]
        cogs_sum = sum(line_cogs) or 1.0
        order_total = _to_float(cell(olines[0], "order_total"))
        flete = _to_float(cell(olines[0], "flete"))
        flete_dev = _to_float(cell(olines[0], "flete_devol"))
        st = _status_bucket(cell(olines[0], "estatus"))
        shop_id = str(cell(olines[0], "shop_order_id") or "").strip()
        plat = shop.platform_for(shop_id) if do_attribution else "unattributed"
        fecha_guia = _parse_fecha(cell(olines[0], "fecha_guia"))
        estatus_norm = _norm(cell(olines[0], "estatus"))
        is_stale = (estatus_norm in EARLY_STUCK_STATUSES
                    and fecha_guia is not None
                    and (today - fecha_guia).days > STALE_PENDING_DAYS)
        if is_stale:
            fl = olines[0]
            stale_orders_collected.append({
                "ID Dropi": cell(fl, "order_id"),
                "Fecha pedido": str(cell(fl, "fecha") or ""),
                "Fecha guía": fecha_guia.strftime("%d/%m/%Y") if fecha_guia else "",
                "Días desde guía": (today - fecha_guia).days if fecha_guia else "",
                "Estatus": cell(fl, "estatus"),
                "Producto": cell(fl, "product_title"),
                "SKU": cell(fl, "sku"),
                "Cantidad": cell(fl, "cantidad"),
                "Total orden": cell(fl, "order_total"),
                "Cliente": cell(fl, "cliente"),
                "Teléfono": cell(fl, "telefono"),
                "Ciudad": cell(fl, "ciudad"),
                "Departamento": cell(fl, "departamento"),
                "Dirección": cell(fl, "direccion"),
                "Transportadora": cell(fl, "transportadora"),
                "Número guía": cell(fl, "numero_guia"),
                "Novedad": cell(fl, "novedad"),
                "Último movimiento": cell(fl, "ultimo_mov"),
                "Fecha último mov.": str(cell(fl, "fecha_ultimo_mov") or ""),
                "ID orden Shopify": shop_id,
                "Notas": cell(fl, "notas"),
            })

        for li, l in enumerate(olines):
            pid = str(cell(l, "product_id") or "").strip() or _norm(
                cell(l, "product_title"))
            b = bucket(pid)
            b["product_title"] = b["product_title"] or str(
                cell(l, "product_title") or "").strip()
            b["sku"] = b["sku"] or str(cell(l, "sku") or "").strip()

            qty = _to_int(cell(l, "cantidad")) or 1
            l_cogs = line_cogs[li]
            # revenue de la línea = prorrateo del total por share de COGS
            share = (line_cogs[li] / cogs_sum) if n_lines > 1 else 1.0
            l_rev = order_total * share
            l_flete = flete * share
            l_flete_dev = flete_dev * share

            b["units"] += qty
            b["revenue_gross"] += l_rev
            b["cogs_gross"] += l_cogs
            b["orders"] += 1
            pu = _to_float(cell(l, "prov_unit"))
            if pu:
                b["prov_unit_sum"] += pu
                b["prov_unit_n"] += 1

            if st == "delivered":
                b["d"] += 1
                b["rev_real"] += l_rev
                b["cogs_real"] += l_cogs
            elif st == "returned":
                b["ret"] += 1
            elif st == "cancelled":
                b["c"] += 1
            else:
                b["pend"] += 1
                if is_stale:
                    b["pend_stale"] += 1

            # Flete: outbound en toda orden confirmada (no cancelada);
            # devolución suma su flete de retorno.
            if st != "cancelled":
                b["ship"] += l_flete
            if st == "returned":
                b["ship"] += l_flete_dev

            # Atribución por plataforma
            if plat == "facebook":
                b["of"] += 1; b["rf"] += l_rev; b["uf"] += qty
            elif plat == "tiktok":
                b["ot"] += 1; b["rt"] += l_rev; b["ut"] += qty
            else:
                b["ou"] += 1; b["ru"] += l_rev; b["uu"] += qty

    # ── Mapeo Shopify product_id ↔ Dropi product_id (vía órdenes joineadas) ──
    # Lo usamos después para agregar Shopify-side por producto Dropi.
    from difflib import SequenceMatcher
    sp_to_dp: dict[str, str] = {}
    for oid, olines in orders.items():
        shop_id = str(cell(olines[0], "shop_order_id") or "").strip()
        if not shop_id or shop_id not in shop.orders:
            continue
        sho = shop.orders[shop_id]
        if len(olines) == 1 and len(sho.line_items) == 1:
            dpid = str(cell(olines[0], "product_id") or "").strip()
            spid = sho.line_items[0].product_id
            if dpid and spid:
                sp_to_dp[spid] = dpid
        else:
            for li in sho.line_items:
                best = ("", 0.0)
                for dl in olines:
                    dtitle = str(cell(dl, "product_title") or "").strip()
                    ratio = SequenceMatcher(None, _norm(li.title),
                                            _norm(dtitle)).ratio()
                    if ratio > best[1]:
                        best = (str(cell(dl, "product_id") or "").strip(),
                                ratio)
                if best[1] > 0.6 and best[0]:
                    sp_to_dp[li.product_id] = best[0]

    # Agregar Shopify-side por producto Dropi + bucket "Shopify sin Dropi".
    shop_per_product: dict[str, dict] = {}
    shop_no_dropi = {"orders": 0, "units": 0, "revenue": 0.0}
    for sho in shop.orders.values():
        dpids_in_order: set[str] = set()
        has_unmapped = False
        for li in sho.line_items:
            dpid = sp_to_dp.get(li.product_id)
            if dpid:
                d = shop_per_product.setdefault(dpid,
                    {"orders": 0, "units": 0, "revenue": 0.0})
                d["units"] += li.quantity
                d["revenue"] += li.price * li.quantity
                dpids_in_order.add(dpid)
            else:
                has_unmapped = True
                shop_no_dropi["units"] += li.quantity
                shop_no_dropi["revenue"] += li.price * li.quantity
        for dpid in dpids_in_order:
            shop_per_product[dpid]["orders"] += 1
        if has_unmapped and not dpids_in_order:
            shop_no_dropi["orders"] += 1

    if shop_no_dropi["orders"] > 0:
        result.warnings.append(
            f"Funnel: {shop_no_dropi['orders']} órdenes Shopify "
            f"(${shop_no_dropi['revenue']:,.0f}) no se sincronizaron a "
            f"Dropi — ahí se cae venta antes del fulfillment.")

    # ── Atribución push-from-campaigns: para cada campaña Meta del rango,
    #    encontrar a qué producto Dropi mapea (id exacto > nombre fuzzy).
    products_list = [(pid, b["product_title"]) for pid, b in agg.items()]
    spend_by_product, attribution_log = ad_spend.attribute_campaigns_to_products(
        meta, products_list)
    result.attribution_log = attribution_log
    unmatched_spend = sum(e["spend"] for e in attribution_log
                          if e["via"] == "unmatched")
    if unmatched_spend > 0:
        result.warnings.append(
            f"${unmatched_spend:,.0f} de spend Meta no matchearon ningún "
            f"producto. Mirá el cruce campaña↔producto para ver cuáles.")

    # ── Materializar filas + fórmulas COD ──
    out_rows: list[ProductReportRow] = []
    total_rev_real_all = sum(b["rev_real"] for b in agg.values()) or 1.0

    for pid, b in agg.items():
        units = b["units"] or 1
        confirmed = b["orders"] - b["c"]
        tasa_conf = _safe_div(confirmed, b["orders"])
        tasa_entr = _safe_div(b["d"], confirmed)

        price = _safe_div(b["revenue_gross"], units)
        cost_item = (_safe_div(b["prov_unit_sum"], b["prov_unit_n"])
                     if b["prov_unit_n"] else _safe_div(b["cogs_gross"], units))

        if use_real:
            revenue_real = b["rev_real"]
            cogs_real = b["cogs_real"]
            shipping_total = b["ship"]
        else:
            # Simulación con sliders del brief
            revenue_real = b["revenue_gross"] * cfg.tasa_entrega
            cogs_real = b["cogs_gross"] * cfg.tasa_entrega
            shipping_total = b["orders"] * cfg.tasa_confirmacion * cfg.costo_envio

        meta_info = spend_by_product.get(pid, {})
        spend_meta = meta_info.get("spend", 0.0) if isinstance(meta_info, dict) else 0.0
        meta_purchases = meta_info.get("purchases", 0) if isinstance(meta_info, dict) else 0
        meta_conv_value = meta_info.get("conv_value", 0.0) if isinstance(meta_info, dict) else 0.0
        roas_meta_rep = _safe_div(meta_conv_value, spend_meta)
        spend_tiktok = 0.0  # TikTok Ads API: fase futura

        sp = shop_per_product.get(pid, {"orders": 0, "units": 0, "revenue": 0.0})
        shopify_orders = sp["orders"]
        shopify_units = sp["units"]
        shopify_revenue = sp["revenue"]

        gastos_share = cfg.gastos_operativos * _safe_div(
            revenue_real, total_rev_real_all)

        profit = (revenue_real - cogs_real - shipping_total
                  - spend_meta - spend_tiktok - gastos_share)
        margin = _safe_div(profit, revenue_real)

        # Benchmark COD-real (lo que necesitás dado tu delivery rate actual)
        gross_margin_cod = revenue_real - cogs_real - shipping_total - gastos_share
        beroas = _safe_div(revenue_real, gross_margin_cod) if gross_margin_cod > 0 else 0.0
        spend_10 = gross_margin_cod - 0.10 * revenue_real
        roas_10 = _safe_div(revenue_real, spend_10) if spend_10 > 0 else 0.0

        # Headline Meta-side: asume 100% entrega y sin envío (lo que Ads Manager
        # muestra como target). Útil para comparar con tu ROAS reportado.
        margin_gross = b["revenue_gross"] - b["cogs_gross"]
        beroas_meta = (_safe_div(b["revenue_gross"], margin_gross)
                       if margin_gross > 0 else 0.0)
        roas_gross_meta = _safe_div(b["revenue_gross"], spend_meta)

        out_rows.append(ProductReportRow(
            product_id=pid,
            product_title=b["product_title"],
            sku=b["sku"],
            price=round(price, 2),
            cost_per_item=round(cost_item, 2),
            units_sold=b["units"],
            revenue_gross=round(b["revenue_gross"], 2),
            cogs_gross=round(b["cogs_gross"], 2),
            orders_count=b["orders"],
            orders_delivered=b["d"],
            orders_cancelled=b["c"],
            orders_returned=b["ret"],
            orders_pending=b["pend"],
            orders_pending_stale=b["pend_stale"],
            tasa_confirmacion_real=round(tasa_conf, 4),
            tasa_entrega_real=round(tasa_entr, 4),
            orders_facebook=b["of"], orders_tiktok=b["ot"],
            orders_unattrib=b["ou"],
            revenue_facebook=round(b["rf"], 2), revenue_tiktok=round(b["rt"], 2),
            revenue_unattrib=round(b["ru"], 2),
            units_facebook=b["uf"], units_tiktok=b["ut"],
            units_unattrib=b["uu"],
            spend_meta=round(spend_meta, 2), spend_tiktok=round(spend_tiktok, 2),
            meta_purchases=meta_purchases,
            meta_conv_value=round(meta_conv_value, 2),
            roas_meta_reported=round(roas_meta_rep, 2),
            shopify_orders=shopify_orders,
            shopify_units=shopify_units,
            shopify_revenue=round(shopify_revenue, 2),
            beroas=round(beroas, 2), roas_10pct=round(roas_10, 2),
            beroas_meta=round(beroas_meta, 2),
            roas_gross_meta=round(roas_gross_meta, 2),
            revenue_real=round(revenue_real, 2), cogs_real=round(cogs_real, 2),
            shipping_total=round(shipping_total, 2),
            profit_neto=round(profit, 2), margin_pct=round(margin, 4),
            roas_real_meta=round(_safe_div(revenue_real, spend_meta), 2),
            roas_real_tiktok=round(_safe_div(revenue_real, spend_tiktok), 2),
        ))

    # Más unidades vendidas arriba (los que mueven la aguja primero)
    out_rows.sort(key=lambda r: -r.units_sold)
    result.rows = out_rows
    result.total_revenue_real = round(sum(r.revenue_real for r in out_rows), 2)
    result.total_cogs_real = round(sum(r.cogs_real for r in out_rows), 2)
    result.total_spend = round(
        sum(r.spend_meta + r.spend_tiktok for r in out_rows), 2)
    result.total_shipping = round(sum(r.shipping_total for r in out_rows), 2)
    result.total_profit_neto = round(sum(r.profit_neto for r in out_rows), 2)
    result.total_pending_stale = sum(r.orders_pending_stale for r in out_rows)
    result.stale_pending_orders = stale_orders_collected
    # Funnel totals: Meta píxel → Shopify → Dropi creadas → Dropi entregadas
    result.total_meta_purchases = sum(r.meta_purchases for r in out_rows)
    result.total_shopify_orders = sum(r.shopify_orders for r in out_rows)
    result.total_shopify_revenue = round(
        sum(r.shopify_revenue for r in out_rows), 2)
    result.total_dropi_created = sum(r.orders_count for r in out_rows)
    result.total_dropi_delivered = sum(r.orders_delivered for r in out_rows)
    result.shopify_no_dropi = {
        "orders": shop_no_dropi["orders"],
        "units": shop_no_dropi["units"],
        "revenue": round(shop_no_dropi["revenue"], 2),
    }
    n(f"Análisis listo: {len(out_rows)} productos · "
      f"profit neto ${result.total_profit_neto:,.0f} · "
      f"funnel Meta {result.total_meta_purchases} → Shopify "
      f"{result.total_shopify_orders} → Dropi "
      f"{result.total_dropi_created}→{result.total_dropi_delivered}")
    return result
